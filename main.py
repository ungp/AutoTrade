import autotrade_tool
import json
import numpy
import openpyxl
import os
import sys
import time
import tensorflow as tf

def main(argv):
    print("Initialization...")
    timestep = 60
    timepacket = 900
    timepredict = 300
    lengthpredict = 300
    ndata = int(timepacket/timestep)
    
    X = tf.placeholder(dtype = tf.float32, shape = [None, ndata, 100, 2])
    Y = tf.placeholder(dtype = tf.float32, shape = [None])

    # Model architecture parameters
    neurons = [ndata*100*2, 1024, 512, 256, 128, 1]

    # Initializers
    sigma = 1
    weight_initializer = tf.variance_scaling_initializer(mode="fan_avg", distribution="uniform", scale=sigma)
    bias_initializer = tf.zeros_initializer()

    # Layers
    hidden = tf.contrib.layers.flatten(X)
    n = 0
    while n < len(neurons) - 1:
        W_hidden = tf.Variable(weight_initializer([neurons[n], neurons[n+1]]))
        bias_hidden = tf.Variable(bias_initializer([neurons[n+1]]))
        if n < len(neurons) - 2:
            hidden = tf.nn.relu(tf.add(tf.matmul(hidden, W_hidden), bias_hidden))
        n = n + 1
    
    # Output layer
    out = tf.transpose(tf.add(tf.matmul(hidden, W_hidden), bias_hidden))
    
    # Cost function
    cost = tf.reduce_mean(tf.squared_difference(out, Y))
    
    # Optimizer
    opt = tf.train.AdamOptimizer().minimize(cost)
    
    # Number of epochs and batch size
    epochs = 1
    batch_size = 32
    
    # Association between currency pair and data columns
    datakeys = {
        "USDT_BTC": ["B", "L"],
        "USDT_ETH": ["C", "M"],
        "USDT_XRP": ["D", "N"],
        "USDT_ETC": ["E", "O"],
        "USDT_BCH": ["F", "P"],
        "USDT_STR": ["G", "Q"],
        "USDT_LTC": ["H", "R"],
        "USDT_XMR": ["I", "S"],
        "USDT_ZEC": ["J", "T"],
        "USDT_DASH": ["K", "U"]
        }
    
    # Start AI
    sess = tf.Session()
    sess.run(tf.global_variables_initializer())

    for e in range(epochs):
        print("Epoch: " + str(e + 1) + "/" + str(epochs))
        
        filenames = numpy.array(os.listdir("data/train"))
        filenames = filenames[numpy.random.permutation(numpy.arange(len(filenames)))]
        for filename in filenames:
            print("Learning from: " + filename + "...")
            sheet = openpyxl.load_workbook("data/train/" + filename)['Sheet1']
            
            for datakey in datakeys:
                datacols = datakeys[datakey]

                X_train = []
                Y_train = []
                i = 1
                imax = sheet.max_row - ((ndata - 1) + int(timepredict/60) + int(lengthpredict/60) - 1)
                while i <= imax:
                    XY_pair = autotrade_tool.data_predict(ndata, sheet, datacols, i, timepredict, lengthpredict)
                    X_train.append(XY_pair[0])
                    Y_train.append(XY_pair[1])
                    autotrade_tool.progress(i, imax, datakey)
                    i = i + 1
                X_train = numpy.array(X_train)
                Y_train = numpy.array(Y_train)
                print("\r")
                
                # Mini-batching
                shuffle_indices = numpy.random.permutation(numpy.arange(len(Y_train)))
                X_train = X_train[shuffle_indices]
                Y_train = Y_train[shuffle_indices]
                for i in range(0, len(Y_train) // batch_size):
                    start = i * batch_size
                    X_batch = X_train[start:start + batch_size]
                    Y_batch = Y_train[start:start + batch_size]
                    # Run optimizer with batch
                    sess.run(opt, feed_dict={X: X_batch, Y: Y_batch})
            
        # Tests after training
        for filename in os.listdir("data/test"):
            testsheet = openpyxl.load_workbook('data/test/' + filename)['Sheet1']
            correct_sign = 0
            err_avg = 0.0
            out_avg = 0.0
            
            j = 0
            for datakey in datakeys:
                datacols = datakeys[datakey]
                i = 1
                imax = testsheet.max_row - ((ndata - 1) + int(timepredict/60) + int(lengthpredict/60) - 1)
                while i <= imax:
                    XY_pair = autotrade_tool.data_predict(ndata, testsheet, datacols, i, timepredict, lengthpredict)
                    X_test = XY_pair[0]
                    Y_test = XY_pair[1]
                    Y_pred = sess.run(out, feed_dict={X: [X_test]})[0][0]
                    if (Y_pred*Y_test > 0):
                        correct_sign = correct_sign + 1
                    err_avg = err_avg + numpy.absolute(Y_pred - Y_test)
                    out_avg = out_avg + numpy.absolute(Y_test)
                    autotrade_tool.progress(i + j * imax, imax * len(datakeys), "Testing")
                    #print(str(Y_pred) + "   " + str(Y_test))
                    i = i + 1
                j = j + 1
            
            correct_sign = round(100.0 * float(correct_sign) / (float(imax * len(datakeys))), 2)
            err_avg = round(err_avg / float(imax * len(datakeys)), 2)
            out_avg = round(out_avg / float(imax * len(datakeys)), 2)
            print("\r")
            print("Correct sign: " + str(correct_sign) + "%")
            print("Average prediction error: " + str(err_avg) + " / " + str(out_avg))
            print("\r")

if __name__ == "__main__":
    main(sys.argv[1:])
